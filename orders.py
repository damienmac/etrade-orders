import time
import datetime
import os
import re
from decimal import Decimal
from typing import Tuple
import pyetrade
import pandas as pd

# Use current date range (two years ago to today)
today = datetime.datetime.now()
two_years_ago = today - datetime.timedelta(days=365*2)

# Format dates as datetime, not MMDDYYYY
from_date = two_years_ago
to_date = today

action_map = {
    "BUY_OPEN": "Buy Open",
    "BUY_CLOSE": "Buy Close",
    "SELL_OPEN": "Sell Open",
    "SELL_CLOSE": "Sell Close",
    "BUY": "Buy",
    "SELL": "Sell",
}

def fetch_executed_orders(etrade_order: pyetrade.order.ETradeOrder,
                          account_id_key: str,
                          from_dt: datetime.datetime,
                          to_dt: datetime.datetime,
                          action_mapping: dict) -> Tuple[list, list]:
    """
    Fetches EXECUTED orders from E*TRADE and separates them into opens and closes.

    :param etrade_order: The ETradeOrder API object.
    :param account_id_key: The account identifier key.
    :param from_dt: The earliest date to include in the date range.
    :param to_dt: The latest date to include in the date range.
    :param action_mapping: A dictionary mapping API actions to display actions.
    :return: A tuple containing (opens, closes) lists.
    """
    opens = []
    closes = []
    done = False
    marker = 0
    while not done:
        order_response = (etrade_order.list_orders(account_id_key,
                                                  marker=marker,
                                                  count=100,
                                                  from_date=from_dt,
                                                  to_date=to_dt)
                          .get("OrdersResponse", {}))
        marker = order_response.get("marker")
        done = not bool(marker)
        for order in order_response.get("Order", []):
            order_id = order.get("orderId", "!NO ORDER ID")
            for detail in order.get("OrderDetail", []):
                status = detail.get("status", "!NO STATUS")
                if status != "EXECUTED":
                    continue
                executed_time = detail.get("executedTime", "!NO EXECUTED TIME")
                # Convert to local time
                local_time = time.localtime(executed_time / 1_000)
                # Format the local time
                formatted_time = time.strftime("%m/%d/%Y", local_time)

                instruments = detail.get("Instrument", [])
                for instrument in instruments:
                    symbol = instrument.get("symbolDescription", "!NO SYMBOL")
                    action = instrument.get("orderAction", "!NO ACTION")
                    action = action_mapping.get(action, "!NO ACTION")
                    quantity = int(instrument.get("filledQuantity", "!NO QUANTITY"))
                    price = Decimal(str(instrument.get("averageExecutionPrice", "0.00")))
                    if action in {"Buy Open", "Buy Close"}:
                        total_in = 0
                        total_out = (price * 100) * quantity * -1
                    elif action == "Buy":
                        total_in = 0
                        total_out = price * quantity * -1
                    elif action in {"Sell Open", "Sell Close"}:
                        total_in = (price * 100) * quantity
                        total_out = 0
                    elif action == "Sell":
                        total_in = price * quantity
                        total_out = 0
                    else:
                        total_in = 0
                        total_out = 0
                    # print(f"{symbol},  {formatted_time}, {action},  {quantity},  {price}")
                    row = {
                        "symbol": symbol,
                        "date": formatted_time,
                        "epoch": executed_time,
                        "action": action,
                        "quantity": quantity,
                        "price": price,
                        "total_in": total_in,
                        "total_out": total_out,
                        "order_id": order_id,
                    }
                    if "Close" in action or "Sell" == action:
                        closes.append(row)
                        # print(f"CLOSES <-- {row}")
                    else:
                        opens.append(row)
                        # print(f"OPENS <-- {row}")
    return opens, closes


def parse_expiration_date(symbol: str) -> datetime.datetime:
    """Parses expiration date from a human-readable option symbol description."""
    try:
        if ' Call' not in symbol and ' Put' not in symbol:
            return None
        parts = symbol.split(' ')
        # Symbol example: "AAPL Apr 17 '25 $200 Call"
        # parts: ["AAPL", "Apr", "17", "'25", "$200", "Call"]
        if len(parts) < 4:
            return None
        # We assume the month and day are at indices 1 and 2, and year at index 3
        month_str = parts[1]
        day_str = parts[2]
        year_str = parts[3].strip("'")
        date_str = f"{month_str} {day_str} {year_str}"
        return datetime.datetime.strptime(date_str, "%b %d %y")
    except (ValueError, IndexError):
        return None


def parse_option_details(symbol: str) -> dict:
    """Parses key details from a human-readable option symbol description."""
    if not symbol:
        return None

    match = re.match(
        r"^(?P<ticker>[A-Z][A-Z0-9\.\-]*)\s+[A-Za-z]{3}\s+\d{1,2}\s+'?\d{2}\s+\$(?P<strike>\d+(?:\.\d+)?)\s+(?P<option_type>Call|Put)\b",
        symbol
    )
    if not match:
        return None

    try:
        strike = Decimal(match.group("strike"))
    except Exception:
        return None

    return {
        "ticker": match.group("ticker"),
        "strike": strike,
        "option_type": match.group("option_type")
    }


def parse_mmddyyyy(date_str: str):
    if not date_str:
        return None
    try:
        if isinstance(date_str, datetime.datetime):
            return date_str.date()
        if isinstance(date_str, datetime.date):
            return date_str
        value = str(date_str).strip()
        if ' ' in value:
            value = value.split(' ')[0]

        for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
            try:
                return datetime.datetime.strptime(value, fmt).date()
            except Exception:
                pass

        return datetime.date.fromisoformat(value)
    except Exception:
        return None


def get_leg_date(leg: dict):
    if not leg:
        return None

    parsed_date = parse_mmddyyyy(leg.get("date"))
    if parsed_date:
        return parsed_date

    epoch = leg.get("epoch")
    if epoch is None:
        return None

    try:
        epoch_int = int(epoch)
        if epoch_int > 10 ** 11:
            epoch_int = epoch_int / 1000
        return datetime.datetime.fromtimestamp(epoch_int).date()
    except Exception:
        return None


def close_on_or_after_option_expiration(option_symbol: str, close_leg: dict) -> bool:
    expiration_dt = parse_expiration_date(option_symbol)
    close_date = get_leg_date(close_leg)
    if not expiration_dt or not close_date:
        return False
    return close_date >= expiration_dt.date()


def leg_sort_value_ms(leg: dict) -> int:
    if not leg:
        return -1
    epoch = leg.get("epoch")
    if epoch is not None:
        try:
            return int(epoch)
        except Exception:
            pass

    parsed_date = parse_mmddyyyy(leg.get("date"))
    if parsed_date:
        return int(datetime.datetime(parsed_date.year, parsed_date.month, parsed_date.day).timestamp() * 1000)
    return -1


def leg_distance_ms(first_leg: dict, second_leg: dict) -> int:
    first_ms = leg_sort_value_ms(first_leg)
    second_ms = leg_sort_value_ms(second_leg)
    if first_ms < 0 or second_ms < 0:
        return 10 ** 18
    return abs(first_ms - second_ms)


def ticker_hint_matches(stock_symbol: str, ticker: str) -> bool:
    if not stock_symbol or not ticker:
        return False

    stock_symbol_upper = stock_symbol.upper()
    ticker_upper = ticker.upper()
    if f"({ticker_upper})" in stock_symbol_upper:
        return True

    tokens = re.findall(r"[A-Z]+", stock_symbol_upper)
    return ticker_upper in tokens


def link_short_put_assignments(combined: list) -> dict:
    """
    Links short-put assignment candidates to related stock buy/sell legs.

    Returns a mapping by combined-entry index:
    {
        put_entry_idx: {
            "status": "ASSIGNED_LINKED"|"ASSIGNED_AMBIGUOUS"|"ASSIGNED_UNRESOLVED",
            "buy_entry_idx": int|None,
            "sell_legs": [close_leg, ...]
        }
    }
    """
    assignment_links = {}
    used_buy_entry_indices = set()
    used_close_only_sell_entry_indices = set()

    stock_buy_entries = []
    close_only_stock_sell_entries = []

    for idx, entry in enumerate(combined):
        symbol = entry.get("symbol", "")
        if " Put" in symbol or " Call" in symbol:
            continue

        opening = entry.get("open")
        closing = entry.get("close")

        if opening and opening.get("action") == "Buy":
            stock_buy_entries.append((idx, entry))

        if (not opening) and closing and closing.get("action") == "Sell":
            close_only_stock_sell_entries.append((idx, entry))

    seven_days_ms = 7 * 24 * 60 * 60 * 1000
    six_hours_ms = 6 * 60 * 60 * 1000
    one_day_ms = 24 * 60 * 60 * 1000
    three_days_ms = 3 * one_day_ms

    for put_idx, put_entry in enumerate(combined):
        opening = put_entry.get("open")
        closing = put_entry.get("close")

        if not opening or not closing:
            continue
        if opening.get("action") != "Sell Open":
            continue
        if " Put" not in opening.get("symbol", ""):
            continue
        if closing.get("action") != "Buy Close":
            continue
        if closing.get("is_expired"):
            continue

        close_price = closing.get("price")
        if close_price is None:
            continue

        try:
            close_price_decimal = Decimal(str(close_price))
        except Exception:
            continue

        if close_price_decimal != Decimal("0.00"):
            continue

        likely_expired_close = close_on_or_after_option_expiration(opening.get("symbol", ""), closing)

        option_details = parse_option_details(opening.get("symbol", ""))
        expected_share_qty = int(opening.get("quantity", 0)) * 100

        if not option_details or expected_share_qty <= 0:
            if likely_expired_close:
                continue
            assignment_links[put_idx] = {
                "status": "ASSIGNED_UNRESOLVED",
                "buy_entry_idx": None,
                "sell_legs": []
            }
            continue

        strike = option_details["strike"]
        ticker = option_details["ticker"]

        candidates = []
        for buy_idx, buy_entry in stock_buy_entries:
            if buy_idx in used_buy_entry_indices:
                continue

            buy_open = buy_entry.get("open")
            if not buy_open:
                continue

            if int(buy_open.get("quantity", 0)) != expected_share_qty:
                continue

            buy_price = buy_open.get("price")
            if buy_price is None or abs(Decimal(str(buy_price)) - strike) > Decimal("0.01"):
                continue

            distance_ms = leg_distance_ms(closing, buy_open)
            if distance_ms > seven_days_ms:
                continue

            score = 0
            if distance_ms <= six_hours_ms:
                score += 4
            elif distance_ms <= one_day_ms:
                score += 3
            elif distance_ms <= three_days_ms:
                score += 2
            else:
                score += 1

            if ticker_hint_matches(buy_open.get("symbol", buy_entry.get("symbol", "")), ticker):
                score += 2

            candidates.append((score, distance_ms, buy_idx))

        if not candidates:
            if likely_expired_close:
                continue
            assignment_links[put_idx] = {
                "status": "ASSIGNED_UNRESOLVED",
                "buy_entry_idx": None,
                "sell_legs": []
            }
            continue

        candidates.sort(key=lambda item: (-item[0], item[1], item[2]))
        best = candidates[0]

        if len(candidates) > 1 and candidates[1][0] == best[0] and candidates[1][1] == best[1]:
            assignment_links[put_idx] = {
                "status": "ASSIGNED_AMBIGUOUS",
                "buy_entry_idx": None,
                "sell_legs": []
            }
            continue

        linked_buy_idx = best[2]
        linked_buy_entry = combined[linked_buy_idx]
        used_buy_entry_indices.add(linked_buy_idx)

        linked_sell_legs = []
        linked_buy_close = linked_buy_entry.get("close")
        matched_sell_qty = 0

        if linked_buy_close and linked_buy_close.get("action") == "Sell":
            linked_sell_legs.append(linked_buy_close)
            matched_sell_qty += int(linked_buy_close.get("quantity", 0) or 0)

        remaining_qty = expected_share_qty - matched_sell_qty
        if remaining_qty > 0:
            buy_symbol = linked_buy_entry.get("symbol")
            buy_open_ms = leg_sort_value_ms(linked_buy_entry.get("open"))

            additional_candidates = []
            for sell_idx, sell_entry in close_only_stock_sell_entries:
                if sell_idx in used_close_only_sell_entry_indices:
                    continue
                if sell_entry.get("symbol") != buy_symbol:
                    continue

                close_leg = sell_entry.get("close")
                if not close_leg or close_leg.get("action") != "Sell":
                    continue

                close_ms = leg_sort_value_ms(close_leg)
                if buy_open_ms >= 0 and close_ms >= 0 and close_ms < buy_open_ms:
                    continue

                additional_candidates.append((close_ms, sell_idx, close_leg))

            additional_candidates.sort(key=lambda item: (item[0], item[1]))
            for _, sell_idx, close_leg in additional_candidates:
                linked_sell_legs.append(close_leg)
                used_close_only_sell_entry_indices.add(sell_idx)
                remaining_qty -= int(close_leg.get("quantity", 0) or 0)
                if remaining_qty <= 0:
                    break

        assignment_links[put_idx] = {
            "status": "ASSIGNED_LINKED",
            "buy_entry_idx": linked_buy_idx,
            "sell_legs": linked_sell_legs
        }

    return assignment_links


def add_expired_worthless_orders(opens: list, closes: list):
    """Adds synthetic closing orders for options that expired worthless."""
    now = datetime.datetime.now()
    
    # Track how many closes we already have for each symbol
    existing_closes_count = {}
    for c in closes:
        symbol = c['symbol']
        existing_closes_count[symbol] = existing_closes_count.get(symbol, 0) + 1
        
    # We want to add synthetic closes for any open that doesn't have a matching close
    # and has expired.
    
    # First, let's count the opens
    opens_count = {}
    for o in opens:
        symbol = o['symbol']
        opens_count[symbol] = opens_count.get(symbol, 0) + 1
        
    # Now, for each symbol, if opens > closes and it's an expired option, 
    # add (opens - closes) synthetic closes.
    for symbol, count in opens_count.items():
        exp_date = parse_expiration_date(symbol)
        if exp_date and exp_date < now:
            num_closes = existing_closes_count.get(symbol, 0)
            if count > num_closes:
                # Get list of opens for this symbol to iterate through
                matching_opens = [o for o in opens if o['symbol'] == symbol]
                
                # Check how many we need to create
                needed = count - num_closes
                # For simplicity, we can just pick the last ones (presumably the most recent)
                # or just any. Since we want to pair them, we should probably use the quantities
                # of the unmatched opens.
                
                # We'll create one synthetic close for each unmatched open
                # But wait, which ones are unmatched? 
                # Match_trades will find any.
                # Let's just create 'needed' synthetic closes using the quantities of some opens.
                for i in range(needed):
                    o_to_use = matching_opens[i]
                    action = o_to_use.get('action', '')
                    if action == "Buy Open":
                        close_action = "Sell Close"
                    elif action == "Sell Open":
                        close_action = "Buy Close"
                    else:
                        continue
                    
                    # E*TRADE timestamps are in milliseconds.
                    epoch = int(exp_date.timestamp() * 1000)
                    
                    # Create a unique-ish ID for synthetic orders to avoid duplicates
                    # Format: SYNTH-[Symbol]-[Date]-[Action]
                    synthetic_id = f"SYNTH-{symbol}-{exp_date.strftime('%Y%m%d')}-{close_action}"

                    synthetic_close = {
                        "symbol": symbol,
                        "date": exp_date.strftime("%m/%d/%Y"),
                        "epoch": epoch,
                        "action": close_action,
                        "quantity": o_to_use.get('quantity'),
                        "price": Decimal("0.00"),
                        "total_in": 0,
                        "total_out": 0,
                        "is_expired": True,
                        "order_id": synthetic_id,
                    }
                    closes.append(synthetic_close)


def match_trades(opens: list, closes: list) -> list:
    """
    Matches opening and closing trades by symbol.

    :param opens: List of opening trades.
    :param closes: List of closing trades.
    :return: A list of matched trade dictionaries.
    """
    # I want the oldest first
    closes.reverse()
    opens.reverse()

    combined = []
    for closing in closes:
        matched = False
        for opening in opens:
            if closing['symbol'] == opening['symbol']:
                match = {
                    "symbol": closing['symbol'],  # for sorting
                    "epoch": closing['epoch'],  # for sorting
                    "open": opening,
                    "close": closing
                }
                combined.append(match)
                # print(f"COMBINED1 <-- {match}")
                matched = True
                opens.remove(opening)
                break
        if not matched:
            match = {
                "symbol": closing['symbol'],
                "epoch": closing['epoch'],
                "open": None,
                "close": closing
            }
            combined.append(match)
            # print(f"COMBINED2 <-- {match}")

    # see what's left in opens that had no closes
    for opening in opens:
        match = {
            "symbol": opening['symbol'],
            "epoch": opening['epoch'],
            "open": opening,
            "close": None
        }
        combined.append(match)
        # print(f"COMBINED3 <-- {match}")
    return combined


def format_output(combined: list) -> list:
    """
    Formats the matched trades into CSV lines, separating puts into their own section.

    :param combined: List of matched trade dictionaries.
    :return: A list of formatted CSV strings.
    """
    output_lines = []
    
    # Add CSV header
    header = "Symbol,Open Date,Open Action,Open Quantity,Open Price,,Open Total Out,Open Total In,,Close Date,Close Action,Close Quantity,Close Price,,Close Total In,Close Total Out"
    output_lines.append(header)

    puts = []
    others = []
    
    for row in sorted(combined, key=lambda x: (x['symbol'], x['epoch'])):
        o = row['open']
        c = row['close']
        if (o and 'Put' in o.get('symbol')) or (c and 'Put' in c.get('symbol')):
            puts.append(row)
        else:
            others.append(row)

    def format_row(row):
        o = row['open']
        c = row['close']
        if o and c:
            return f"{o.get('symbol')},{o.get('date')},{o.get('action')},{o.get('quantity')},{o.get('price')},,{o.get('total_out')},{o.get('total_in')},,{c.get('date')},{c.get('action')},{c.get('quantity')},{c.get('price')},,{c.get('total_in')},{c.get('total_out')}"
        elif o and not c:
            return f"{o.get('symbol')},{o.get('date')},{o.get('action')},{o.get('quantity')},{o.get('price')},,{o.get('total_out')},{o.get('total_in')},,,,,,"
        elif not o and c:
            return f"{c.get('symbol')},,,,,,,,,{c.get('date')},{c.get('action')},{c.get('quantity')},{c.get('price')},,{c.get('total_in')},{c.get('total_out')}"
        else:
            assert False, f"Empty row? {row}"

    for row in others:
        output_lines.append(format_row(row))
        
    for row in puts:
        output_lines.append(format_row(row))
        
    return output_lines


def write_output(output_lines: list, output_file: str = None):
    """
    Writes the output lines to a file or the console.

    :param output_lines: List of formatted CSV strings.
    :param output_file: The path to the output file (optional).
    """
    if output_file:
        try:
            with open(output_file, 'w') as f:
                for line in output_lines:
                    f.write(line + '\n')
            print(f"Output saved to {output_file}")
        except Exception as e:
            print(f"Error writing to output file: {e}")
            # Fall back to console output
            for line in output_lines:
                print(line)
    else:
        # Print to console
        for line in output_lines:
            print(line)


def load_previous_output(output_file: str) -> list:
    """
    Loads previous trades from an Excel or CSV file to enable 'bringing forward' historical data.
    """
    if not output_file:
        return []

    original_output_file = output_file
    
    # If the output file is still .csv, we'll check it, but also check for Excel files
    if output_file.lower().endswith('.csv'):
        xlsx_file = output_file[:-4] + '.xlsx'
    else:
        xlsx_file = output_file

    directory = os.path.dirname(xlsx_file) or '.'
    base_name = os.path.basename(xlsx_file)
    # Remove extension and date pattern if exists to find the prefix
    # We look for files starting with 'orders_output'
    prefix = base_name.split('.')[0]
    if '_' in prefix:
        prefix = prefix.split('_')[0]

    # Find the most recent Excel file if it doesn't exist exactly as named (e.g. dated files)
    if not os.path.exists(xlsx_file):
        extensions = ['.xlsx', '.xlsm']
        files = [f for f in os.listdir(directory) if f.startswith(prefix) and any(f.endswith(ext) for ext in extensions) and not f.startswith('~$')]
        if files:
            # Sort by modification time to get the latest
            files.sort(key=lambda x: os.path.getmtime(os.path.join(directory, x)), reverse=True)
            xlsx_file = os.path.join(directory, files[0])
            print(f"Loading previous trades from: {xlsx_file}")
        else:
            xlsx_file = None
    else:
        print(f"Loading previous trades from exact file: {xlsx_file}")

    all_history = []

    def normalize_column_name(col_name: str) -> str:
        return re.sub(r'\s+', ' ', str(col_name)).strip()

    def parse_quantity(qty_value):
        if pd.isna(qty_value):
            return None
        try:
            qty = int(float(str(qty_value).replace(',', '').strip()))
            return qty if qty > 0 else None
        except (TypeError, ValueError):
            return None

    if xlsx_file and os.path.exists(xlsx_file):
        try:
            xls = pd.ExcelFile(xlsx_file)
            for sheet_name in xls.sheet_names:
                if sheet_name == 'Dashboard':
                    continue
                
                df = pd.read_excel(xls, sheet_name=sheet_name)
                # Filter out completely empty rows
                df = df.dropna(how='all')
                if df.empty:
                    continue

                # Normalize whitespace so headers like "Open\nQuantity" map to "Open Quantity"
                df.rename(columns={col: normalize_column_name(col) for col in df.columns}, inplace=True)

                # Standardize columns (some might be missing in older versions)
                required_cols = [
                    'Symbol', 'Open Date', 'Open Action', 'Open Quantity', 'Open Price', 
                    'Open Total Out', 'Open Total In', 'Close Date', 'Close Action', 
                    'Close Quantity', 'Close Price', 'Close Total In', 'Close Total Out'
                ]
                
                # Check for legacy column names in single-sheet formats
                if 'Symbol' not in df.columns:
                    # Try to find a column that looks like Symbol
                    for col in df.columns:
                        if 'symbol' in str(col).lower():
                            df.rename(columns={col: 'Symbol'}, inplace=True)
                            break
                
                if 'Symbol' not in df.columns:
                    continue

                for col in required_cols:
                    if col not in df.columns:
                        df[col] = None
                
                # Map back to internal 'combined' structure
                for _, row in df.iterrows():
                    # Skip if symbol is missing (likely an empty or header row)
                    if pd.isna(row['Symbol']):
                        continue

                    # Skip synthetic assignment rows from previously generated short-put tabs.
                    # They are reporting-only duplicates of stock legs and should not be reloaded
                    # as historical source trades.
                    strategy_event = str(row.get('Strategy Event')).strip().upper() if pd.notna(row.get('Strategy Event')) else ''
                    if strategy_event in {'ASSIGNMENT BUY', 'ASSIGNMENT SELL'}:
                        continue

                    # Reconstruct 'open' part
                    opening = None
                    open_qty = parse_quantity(row['Open Quantity'])
                    if pd.notna(row['Open Date']) and open_qty is not None:
                        opening = {
                            "symbol": str(row['Symbol']),
                            "date": str(row['Open Date']),
                            "action": str(row['Open Action']),
                            "quantity": open_qty,
                            "price": Decimal(str(row['Open Price'])) if pd.notna(row['Open Price']) else None,
                            "total_in": Decimal(str(row['Open Total In'])) if pd.notna(row['Open Total In']) else 0,
                            "total_out": Decimal(str(row['Open Total Out'])) if pd.notna(row['Open Total Out']) else 0,
                            "order_id": row.get('Open Order ID') if pd.notna(row.get('Open Order ID')) else None,
                        }
                        # Try to reconstruct epoch from date
                        parsed_open_date = parse_mmddyyyy(row['Open Date'])
                        if parsed_open_date:
                            opening["epoch"] = int(datetime.datetime(parsed_open_date.year, parsed_open_date.month, parsed_open_date.day).timestamp() * 1000)
                        else:
                            opening["epoch"] = None
                    
                    # Reconstruct 'close' part
                    closing = None
                    close_qty = parse_quantity(row['Close Quantity'])
                    if pd.notna(row['Close Date']) and close_qty is not None:
                        closing = {
                            "symbol": str(row['Symbol']),
                            "date": str(row['Close Date']),
                            "action": str(row['Close Action']),
                            "quantity": close_qty,
                            "price": Decimal(str(row['Close Price'])) if pd.notna(row['Close Price']) else None,
                            "total_in": Decimal(str(row['Close Total In'])) if pd.notna(row['Close Total In']) else 0,
                            "total_out": Decimal(str(row['Close Total Out'])) if pd.notna(row['Close Total Out']) else 0,
                            "is_expired": row.get('EXPIRED') == "EXPIRED",
                            "order_id": row.get('Close Order ID') if pd.notna(row.get('Close Order ID')) else None,
                        }
                        # Try to reconstruct epoch from date
                        parsed_close_date = parse_mmddyyyy(row['Close Date'])
                        if parsed_close_date:
                            closing["epoch"] = int(datetime.datetime(parsed_close_date.year, parsed_close_date.month, parsed_close_date.day).timestamp() * 1000)
                        else:
                            closing["epoch"] = None

                    trade = {
                        "symbol": str(row['Symbol']),
                        "epoch": closing.get("epoch") if (closing and closing.get("epoch") is not None) else (opening.get("epoch") if (opening and opening.get("epoch") is not None) else 0),
                        "open": opening,
                        "close": closing
                    }
                    if opening or closing:
                        all_history.append(trade)
        except Exception as e:
            print(f"Warning: Could not load previous Excel output file: {e}")

    # ALSO load from legacy CSV if it exists to ensure we don't miss anything
    if os.path.exists('orders_output.csv'):
        try:
            print(f"Merging legacy trades from CSV: orders_output.csv")
            df = pd.read_csv('orders_output.csv')
            # Filter out completely empty rows
            df = df.dropna(how='all')
            df.rename(columns={col: normalize_column_name(col) for col in df.columns}, inplace=True)
            
            # Map back to internal 'combined' structure
            for _, row in df.iterrows():
                # Skip if symbol is missing
                if pd.isna(row['Symbol']):
                    continue

                strategy_event = str(row.get('Strategy Event')).strip().upper() if pd.notna(row.get('Strategy Event')) else ''
                if strategy_event in {'ASSIGNMENT BUY', 'ASSIGNMENT SELL'}:
                    continue

                # Reconstruct 'open' part
                opening = None
                open_qty = parse_quantity(row['Open Quantity'])
                if pd.notna(row['Open Date']) and open_qty is not None:
                    opening = {
                        "symbol": str(row['Symbol']),
                        "date": str(row['Open Date']),
                        "action": str(row['Open Action']),
                        "quantity": open_qty,
                        "price": Decimal(str(row['Open Price'])) if pd.notna(row['Open Price']) else None,
                        "total_in": Decimal(str(row['Open Total In'])) if pd.notna(row['Open Total In']) else 0,
                        "total_out": Decimal(str(row['Open Total Out'])) if pd.notna(row['Open Total Out']) else 0,
                    }
                    parsed_open_date = parse_mmddyyyy(row['Open Date'])
                    if parsed_open_date:
                        opening["epoch"] = int(datetime.datetime(parsed_open_date.year, parsed_open_date.month, parsed_open_date.day).timestamp() * 1000)
                    else:
                        opening["epoch"] = None
                
                # Reconstruct 'close' part
                closing = None
                close_qty = parse_quantity(row['Close Quantity'])
                if pd.notna(row['Close Date']) and close_qty is not None:
                    closing = {
                        "symbol": str(row['Symbol']),
                        "date": str(row['Close Date']),
                        "action": str(row['Close Action']),
                        "quantity": close_qty,
                        "price": Decimal(str(row['Close Price'])) if pd.notna(row['Close Price']) else None,
                        "total_in": Decimal(str(row['Close Total In'])) if pd.notna(row['Close Total In']) else 0,
                        "total_out": Decimal(str(row['Close Total Out'])) if pd.notna(row['Close Total Out']) else 0,
                    }
                    parsed_close_date = parse_mmddyyyy(row['Close Date'])
                    if parsed_close_date:
                        closing["epoch"] = int(datetime.datetime(parsed_close_date.year, parsed_close_date.month, parsed_close_date.day).timestamp() * 1000)
                    else:
                        closing["epoch"] = None

                trade = {
                    "symbol": str(row['Symbol']),
                    "epoch": closing.get("epoch") if (closing and closing.get("epoch") is not None) else (opening.get("epoch") if (opening and opening.get("epoch") is not None) else 0),
                    "open": opening,
                    "close": closing
                }
                if opening or closing:
                    all_history.append(trade)
        except Exception as e:
            print(f"Warning: Could not load legacy CSV file: {e}")

    return all_history


def merge_and_deduplicate(old_trades: list, new_trades: list) -> list:
    """
    Merges old and new trades using a hybrid ID and fingerprint approach.
    """
    def normalize_date_for_key(date_value):
        parsed_date = parse_mmddyyyy(date_value)
        if parsed_date:
            return parsed_date.isoformat()

        raw = str(date_value).strip() if date_value is not None else ""
        if ' ' in raw:
            raw = raw.split(' ')[0]
        return raw

    def normalize_price_for_key(price_value):
        if price_value is None:
            return ""
        try:
            price = Decimal(str(price_value))
            return str(price.normalize())
        except Exception:
            return str(price_value)

    def get_fingerprint(trade_leg):
        if not trade_leg:
            return None
        symbol = str(trade_leg.get('symbol', '')).strip()
        date_key = normalize_date_for_key(trade_leg.get('date'))
        action = str(trade_leg.get('action', '')).strip()
        quantity = int(trade_leg.get('quantity', 0) or 0)
        price_key = normalize_price_for_key(trade_leg.get('price'))
        # Fingerprint: Symbol|Date|Action|Quantity|Price
        return f"{symbol}|{date_key}|{action}|{quantity}|{price_key}"

    def get_order_leg_key(trade_leg):
        if not trade_leg:
            return None
        order_id = trade_leg.get('order_id')
        if order_id is None:
            return None

        symbol = str(trade_leg.get('symbol', '')).strip()
        action = str(trade_leg.get('action', '')).strip()
        quantity = int(trade_leg.get('quantity', 0) or 0)
        price_key = normalize_price_for_key(trade_leg.get('price'))
        date_key = normalize_date_for_key(trade_leg.get('date'))
        return f"{order_id}|{symbol}|{action}|{quantity}|{price_key}|{date_key}"

    # We want to keep track of legs (opens and closes) independently to ensure full deduplication
    seen_order_leg_keys = set()
    seen_fingerprints = set()
    
    unique_trades = []
    
    # Process new trades first as they are "fresher" and have Order IDs
    for trade in new_trades + old_trades:
        o = trade['open']
        c = trade['close']
        
        # Determine if this trade is "new" to our list
        # A trade is considered seen if BOTH its legs (if they exist) have been seen
        legs_seen = 0
        legs_count = 0
        
        if o:
            legs_count += 1
            o_id_key = get_order_leg_key(o)
            o_fp = get_fingerprint(o)
            if (o_id_key and o_id_key in seen_order_leg_keys) or (o_fp in seen_fingerprints):
                legs_seen += 1

        if c:
            legs_count += 1
            c_id_key = get_order_leg_key(c)
            c_fp = get_fingerprint(c)
            if (c_id_key and c_id_key in seen_order_leg_keys) or (c_fp in seen_fingerprints):
                legs_seen += 1
                
        if legs_seen < legs_count:
            # At least one leg is new, so we add this trade
            unique_trades.append(trade)
            
            # Mark legs as seen
            if o:
                o_id_key = get_order_leg_key(o)
                if o_id_key:
                    seen_order_leg_keys.add(o_id_key)
                seen_fingerprints.add(get_fingerprint(o))
            if c:
                c_id_key = get_order_leg_key(c)
                if c_id_key:
                    seen_order_leg_keys.add(c_id_key)
                seen_fingerprints.add(get_fingerprint(c))
                
    return unique_trades


def write_excel_output(combined: list, output_file: str):
    """
    Writes the matched trades to an Excel file with data divided across multiple sheets.

    :param combined: List of matched trade dictionaries.
    :param output_file: The path to the output Excel file.
    """
    if not output_file:
        print("No output file specified for Excel output.")
        return

    # If the output file is still .csv, change it to .xlsx and add current date
    if output_file.lower().endswith('.csv'):
        output_file = output_file[:-4]
    elif output_file.lower().endswith('.xlsx'):
        output_file = output_file[:-5]
    
    # Add today's date to filename
    datestr = datetime.datetime.now().strftime("%Y-%m-%d")
    output_file = f"{output_file}_{datestr}.xlsx"

    rows = []
    validation_rows = []
    assignment_links = link_short_put_assignments(combined)
    assignment_linked_buy_indices = {
        info.get("buy_entry_idx")
        for info in assignment_links.values()
        if info.get("status") == "ASSIGNED_LINKED" and info.get("buy_entry_idx") is not None
    }

    def assignment_leg_key(leg: dict):
        if not leg:
            return None
        symbol = str(leg.get('symbol', '')).strip()
        action = str(leg.get('action', '')).strip()
        quantity = int(leg.get('quantity', 0) or 0)
        date_value = parse_mmddyyyy(leg.get('date'))
        date_key = date_value.isoformat() if date_value else str(leg.get('date') or '').split(' ')[0]
        try:
            price_key = str(Decimal(str(leg.get('price'))).normalize()) if leg.get('price') is not None else ""
        except Exception:
            price_key = str(leg.get('price'))
        order_id_key = str(leg.get('order_id') or '')
        return f"{symbol}|{action}|{quantity}|{price_key}|{date_key}|{order_id_key}"

    assignment_linked_sell_keys = {
        assignment_leg_key(sell_leg)
        for info in assignment_links.values()
        if info.get("status") == "ASSIGNED_LINKED"
        for sell_leg in info.get("sell_legs", [])
        if assignment_leg_key(sell_leg)
    }

    entries_with_indices = list(enumerate(combined))

    def parse_to_datetime(date_str):
        if not date_str:
            return None
        try:
            # Handle potential mixed formats or objects
            if isinstance(date_str, (datetime.datetime, datetime.date)):
                return datetime.datetime(date_str.year, date_str.month, date_str.day)
            return datetime.datetime.strptime(date_str, "%m/%d/%Y")
        except Exception:
            return date_str

    def determine_close_year(close_leg):
        if not close_leg:
            return None

        close_epoch = close_leg.get('epoch')
        if close_epoch is not None:
            try:
                close_time = datetime.datetime.fromtimestamp(int(close_epoch) / 1000)
                if close_time.year > 1980:
                    return close_time.year
            except Exception:
                pass

        parsed_date = parse_mmddyyyy(close_leg.get('date'))
        if parsed_date and parsed_date.year > 1980:
            return parsed_date.year
        return None

    def build_row_data(symbol, opening, closing, is_sold_put, close_year, strategy_link_id=None, strategy_event=None, assignment_status=None):
        return {
            "Symbol": symbol,
            "Open Date": parse_to_datetime(opening.get('date')) if opening else None,
            "Open Action": opening.get('action') if opening else None,
            "Open\nQuantity": opening.get('quantity') if opening else None,
            "Open Price": float(opening.get('price')) if opening and opening.get('price') is not None else None,
            "Open Total Out": float(opening.get('total_out')) if opening and opening.get('total_out') is not None else None,
            "Open Total In": float(opening.get('total_in')) if opening and opening.get('total_in') is not None else None,
            "Close Date": parse_to_datetime(closing.get('date')) if closing else None,
            "Close Action": closing.get('action') if closing else None,
            "Close\nQuantity": closing.get('quantity') if closing else None,
            "Close Price": float(closing.get('price')) if closing and closing.get('price') is not None else None,
            "Close Total In": float(closing.get('total_in')) if closing and closing.get('total_in') is not None else None,
            "Close Total Out": float(closing.get('total_out')) if closing and closing.get('total_out') is not None else None,
            "EXPIRED": "EXPIRED" if closing and closing.get('is_expired') else "",
            "Open Order ID": opening.get('order_id') if opening else None,
            "Close Order ID": closing.get('order_id') if closing else None,
            "Strategy Link ID": strategy_link_id,
            "Strategy Event": strategy_event,
            "Assignment Status": assignment_status,
            "_is_sold_put": is_sold_put,
            "_close_year": close_year
        }

    def build_validation_issues(df_all: pd.DataFrame):
        if df_all.empty:
            return pd.DataFrame(columns=list(df_all.columns) + ["ValidationIssueType", "ValidationReason"])

        issues = []
        option_rows = df_all[df_all["Symbol"].astype(str).str.contains(" Call| Put", na=False)].copy()
        option_rows = option_rows[option_rows["Assignment Status"].astype(str) != "ASSIGNED_LINKED"]

        if option_rows.empty:
            return pd.DataFrame(columns=list(df_all.columns) + ["ValidationIssueType", "ValidationReason"])

        for symbol, group in option_rows.groupby("Symbol", dropna=False):
            open_qty_total = group[
                group["Open Action"].isin(["Buy Open", "Sell Open"])
            ]["Open\nQuantity"].fillna(0).sum()
            close_qty_total = group[
                group["Close Action"].isin(["Buy Close", "Sell Close"])
            ]["Close\nQuantity"].fillna(0).sum()

            unmatched_close_qty = int(max(0, close_qty_total - open_qty_total))
            if unmatched_close_qty <= 0:
                continue

            close_only_rows = group[
                group["Close Action"].isin(["Buy Close", "Sell Close"])
                & group["Open Action"].isna()
                & group["Close\nQuantity"].fillna(0).gt(0)
            ].copy()

            if close_only_rows.empty:
                continue

            close_only_rows = close_only_rows.sort_values(by=["Close Date", "Close\nQuantity"], ascending=[True, False])
            expiration_dt = parse_expiration_date(str(symbol))

            for _, candidate in close_only_rows.iterrows():
                if unmatched_close_qty <= 0:
                    break

                candidate_close_date = candidate.get("Close Date")
                if pd.isna(candidate_close_date):
                    continue

                if expiration_dt is None or expiration_dt.date() >= datetime.date.today():
                    continue

                qty = int(candidate.get("Close\nQuantity") or 0)
                if qty <= 0:
                    continue

                issue_row = candidate.to_dict()
                issue_row["ValidationIssueType"] = "historical_orphan_close"
                issue_row["ValidationReason"] = "Close exists without enough matching open quantity; contract expiration is in the past."
                issues.append(issue_row)
                unmatched_close_qty -= qty

        issue_columns = list(df_all.columns) + ["ValidationIssueType", "ValidationReason"]
        if not issues:
            return pd.DataFrame(columns=issue_columns)
        return pd.DataFrame(issues, columns=issue_columns)

    for entry_idx, entry in sorted(entries_with_indices, key=lambda item: (item[1]['symbol'], item[1]['epoch'])):
        o = entry['open']
        c = entry['close']

        if entry_idx in assignment_linked_buy_indices:
            continue

        if (not o) and c and c.get('action') == 'Sell':
            close_key = assignment_leg_key(c)
            if close_key and close_key in assignment_linked_sell_keys:
                continue

        # Determine if it's a "sold put"
        # A sold put is typically an opening transaction with "Sell Open" action and "Put" in symbol
        is_sold_put = False
        if o and 'Put' in o.get('symbol', '') and 'Sell Open' == o.get('action'):
            is_sold_put = True
        elif c and 'Put' in c.get('symbol', '') and not o:
            # If we only have a close, we might not know for sure if it was a sold put 
            # unless we look at the action. But the user defined "sold puts" as a category.
            # Usually SELL_OPEN is the indicator for sold puts.
            # For now, let's stick to the opening action if available.
            pass

        close_year = determine_close_year(c)
        assignment_info = assignment_links.get(entry_idx)

        strategy_link_id = None
        strategy_event = None
        assignment_status = None
        if assignment_info:
            strategy_link_id = f"SPASSIGN-{(o or c).get('order_id') or entry_idx}"
            strategy_event = "SHORT PUT"
            assignment_status = assignment_info.get("status")

        row_data = build_row_data(
            symbol=(o or c).get('symbol'),
            opening=o,
            closing=c,
            is_sold_put=is_sold_put,
            close_year=close_year,
            strategy_link_id=strategy_link_id,
            strategy_event=strategy_event,
            assignment_status=assignment_status
        )
        rows.append(row_data)
        validation_rows.append(row_data.copy())

        if assignment_info and assignment_info.get("status") == "ASSIGNED_LINKED":
            buy_entry_idx = assignment_info.get("buy_entry_idx")
            linked_buy_entry = None
            linked_buy_symbol = None
            if buy_entry_idx is not None:
                linked_buy_entry = combined[buy_entry_idx]
                linked_buy_open = linked_buy_entry.get("open")
                linked_buy_close = linked_buy_entry.get("close")
                linked_buy_symbol = linked_buy_entry.get("symbol")
                assignment_buy_close_year = (
                    determine_close_year(linked_buy_close)
                    or determine_close_year(c)
                    or determine_close_year(linked_buy_open)
                )

                assignment_buy_row = build_row_data(
                    symbol=linked_buy_symbol,
                    opening=linked_buy_open,
                    closing={},
                    is_sold_put=True,
                    close_year=assignment_buy_close_year,
                    strategy_link_id=strategy_link_id,
                    strategy_event="ASSIGNMENT BUY",
                    assignment_status="ASSIGNED_LINKED"
                )
                rows.append(assignment_buy_row)
                validation_rows.append(assignment_buy_row.copy())

            for sell_leg in assignment_info.get("sell_legs", []):
                if not sell_leg:
                    continue
                assignment_sell_row = build_row_data(
                    symbol=sell_leg.get("symbol") or linked_buy_symbol,
                    opening={},
                    closing=sell_leg,
                    is_sold_put=True,
                    close_year=determine_close_year(sell_leg),
                    strategy_link_id=strategy_link_id,
                    strategy_event="ASSIGNMENT SELL",
                    assignment_status="ASSIGNED_LINKED"
                )
                rows.append(assignment_sell_row)
                validation_rows.append(assignment_sell_row.copy())

    # Convert to a DataFrame
    df_raw = pd.DataFrame(rows)
    validation_df_raw = pd.DataFrame(validation_rows)
    
    # Convert date columns to datetime objects so pandas/openpyxl can handle them as dates
    for col in ["Open Date", "Close Date"]:
        df_raw[col] = pd.to_datetime(df_raw[col], errors='coerce').dt.date
        validation_df_raw[col] = pd.to_datetime(validation_df_raw[col], errors='coerce').dt.date

    validation_issues_df = build_validation_issues(validation_df_raw)

    df = df_raw
    
    this_year = datetime.datetime.now().year

    # Partitioning logic - Dynamic years
    # We want sheets for every year present in the data, plus a 'Current or Open' sheet
    all_years = sorted([y for y in df['_close_year'].unique() if pd.notna(y)], reverse=True)
    
    sheets = []
    
    # Always include 'Current or Open' first (it will be Year 2026 if run in 2026, or trades with no close year)
    current_trades = df[(~df['_is_sold_put']) & ((df['_close_year'] == this_year) | (df['_close_year'].isna()))]
    current_puts = df[(df['_is_sold_put']) & ((df['_close_year'] == this_year) | (df['_close_year'].isna()))]
    
    if not current_trades.empty:
        sheets.append((current_trades, "Trades Current or Open"))
    if not current_puts.empty:
        sheets.append((current_puts, "Short Puts Current or Open"))
    
    # Then sheets for each previous year
    for year in all_years:
        if year == this_year:
            continue
        year_trades = df[(~df['_is_sold_put']) & (df['_close_year'] == year)]
        year_puts = df[(df['_is_sold_put']) & (df['_close_year'] == year)]
        if not year_trades.empty:
            sheets.append((year_trades, f"Trades {int(year)}"))
        if not year_puts.empty:
            sheets.append((year_puts, f"Short Puts {int(year)}"))

    # Remove helper columns before writing
    cols_to_drop = ['_is_sold_put', '_close_year']
    
    # Dashboard calculations
    def calculate_summary(df_partition, name):
        if df_partition.empty:
            return {
                "Category": name,
                "Total Trades": 0,
                "Closed Trades": 0,
                "Total P/L": 0.0,
                "Win Rate (Closed)": "N/A"
            }
        # P/L = (Open Total In + Open Total Out) + (Close Total In + Close Total Out)
        pl = (df_partition['Open Total Out'].fillna(0) + 
              df_partition['Open Total In'].fillna(0) + 
              df_partition['Close Total In'].fillna(0) + 
              df_partition['Close Total Out'].fillna(0)).sum()
        
        count = len(df_partition)
        
        # Win rate: only for closed trades
        closed_trades = df_partition[df_partition['Close Date'].notna()]
        if len(closed_trades) > 0:
            trade_pls = (closed_trades['Open Total Out'].fillna(0) + 
                         closed_trades['Open Total In'].fillna(0) + 
                         closed_trades['Close Total In'].fillna(0) + 
                         closed_trades['Close Total Out'].fillna(0))
            wins = (trade_pls > 0).sum()
            win_rate = f"{(wins / len(closed_trades)) * 100:.2f}%"
        else:
            win_rate = "N/A"
            
        return {
            "Category": name,
            "Total Trades": count,
            "Closed Trades": len(closed_trades),
            "Total P/L": round(pl, 2),
            "Win Rate (Closed)": win_rate
        }

    summary_data = [calculate_summary(data, name) for data, name in sheets]
    dashboard_df = pd.DataFrame(summary_data)

    if validation_issues_df.empty:
        validation_summary_df = pd.DataFrame([
            {
                "Validation Issue Type": "None",
                "Rows": 0,
                "Distinct Symbols": 0,
                "Total Close Quantity": 0,
                "Net Cash Impact": 0.0,
                "Gross Cash Moved": 0.0,
            }
        ])
    else:
        validation_issues_df = validation_issues_df.copy()
        for cash_col in ["Open Total In", "Open Total Out", "Close Total In", "Close Total Out"]:
            validation_issues_df[cash_col] = pd.to_numeric(validation_issues_df[cash_col], errors="coerce").fillna(0.0)
        validation_issues_df["_net_cash_impact"] = (
            validation_issues_df["Open Total In"]
            + validation_issues_df["Open Total Out"]
            + validation_issues_df["Close Total In"]
            + validation_issues_df["Close Total Out"]
        )
        validation_issues_df["_gross_cash_moved"] = (
            validation_issues_df["Open Total In"].abs()
            + validation_issues_df["Open Total Out"].abs()
            + validation_issues_df["Close Total In"].abs()
            + validation_issues_df["Close Total Out"].abs()
        )

        validation_summary_df = (
            validation_issues_df.groupby("ValidationIssueType", dropna=False)
            .agg({
                "Symbol": "nunique",
                "Close\nQuantity": "sum",
                "_net_cash_impact": "sum",
                "_gross_cash_moved": "sum",
            })
            .reset_index()
            .rename(columns={
                "ValidationIssueType": "Validation Issue Type",
                "Symbol": "Distinct Symbols",
                "Close\nQuantity": "Total Close Quantity",
                "_net_cash_impact": "Net Cash Impact",
                "_gross_cash_moved": "Gross Cash Moved",
            })
        )
        issue_counts = validation_issues_df["ValidationIssueType"].value_counts(dropna=False).rename_axis(
            "Validation Issue Type"
        ).reset_index(name="Rows")
        validation_summary_df = issue_counts.merge(validation_summary_df, on="Validation Issue Type", how="left")
        validation_summary_df = validation_summary_df.sort_values(by=["Rows", "Total Close Quantity"], ascending=[False, False])

    accounting_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    date_format = 'mm/dd/yyyy'  # This maps to Excel's Short Date in many locales
    from openpyxl.styles import Alignment

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        dashboard_df.to_excel(writer, sheet_name='Dashboard', index=False)
        # Apply formatting to Dashboard
        worksheet = writer.sheets['Dashboard']
        for col_idx, col_name in enumerate(dashboard_df.columns, 1):
            if col_name == "Total P/L":
                for row_idx in range(2, len(dashboard_df) + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.number_format = accounting_format
            
            # Auto-fit column width
            max_length = 0
            column_letter = worksheet.cell(row=1, column=col_idx).column_letter
            # Header length
            header_lines = str(col_name).split('\n')
            max_length = max(max_length, max(len(line) for line in header_lines))
            if len(header_lines) > 1:
                worksheet.cell(row=1, column=col_idx).alignment = Alignment(wrapText=True, horizontal='center', vertical='bottom')

            # Data length
            for row_idx in range(2, len(dashboard_df) + 2):
                cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    val_str = str(cell_value)
                    if col_name == "Total P/L":
                        val_str = "$#,###,###.00" # wider typical currency length
                    max_length = max(max_length, len(val_str))
            
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        validation_count = len(validation_issues_df)
        pointer_row = len(dashboard_df) + 3
        worksheet.cell(row=pointer_row, column=1, value="Validation Issues")
        worksheet.cell(row=pointer_row, column=2, value=validation_count)
        worksheet.cell(
            row=pointer_row,
            column=3,
            value="See 'Validation Issues' and 'Validation Summary' tabs",
        )

        for data, name in sheets:
            df_to_write = data.drop(columns=cols_to_drop)
            df_to_write.to_excel(writer, sheet_name=name, index=False)
            
            # Apply formatting to yearly sheets
            worksheet = writer.sheets[name]
            for col_idx, col_name in enumerate(df_to_write.columns, 1):
                # Price and Total columns
                if "Price" in col_name or "Total" in col_name:
                    for row_idx in range(2, len(df_to_write) + 2):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.number_format = accounting_format
                
                # Date columns
                if "Date" in col_name:
                    for row_idx in range(2, len(df_to_write) + 2):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        # Setting to 'mm-dd-yy' or 'm/d/yy' often maps to the built-in 
                        # 'Short Date' format (format ID 14) in Excel.
                        cell.number_format = 'm/d/yy'
                
                # Auto-fit column width
                max_length = 0
                column_letter = worksheet.cell(row=1, column=col_idx).column_letter
                # Header length
                header_lines = str(col_name).split('\n')
                max_length = max(max_length, max(len(line) for line in header_lines))
                if len(header_lines) > 1:
                    worksheet.cell(row=1, column=col_idx).alignment = Alignment(wrapText=True, horizontal='center', vertical='bottom')

                # Data length
                for row_idx in range(2, len(df_to_write) + 2):
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        # For dates and currency, we might want a bit more padding
                        val_str = str(cell_value)
                        if "Date" in col_name:
                            val_str = "MM/DD/YYYY" # typical date length
                        elif "Price" in col_name or "Total" in col_name:
                            val_str = "$#,###.00" # typical currency length
                        
                        max_length = max(max_length, len(val_str))
                
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        validation_to_write = validation_issues_df.drop(columns=cols_to_drop, errors='ignore')
        validation_to_write.to_excel(writer, sheet_name="Validation Issues", index=False)

        validation_sheet = writer.sheets["Validation Issues"]
        for col_idx, col_name in enumerate(validation_to_write.columns, 1):
            if "Price" in col_name or "Total" in col_name:
                for row_idx in range(2, len(validation_to_write) + 2):
                    cell = validation_sheet.cell(row=row_idx, column=col_idx)
                    cell.number_format = accounting_format

            if "Date" in col_name:
                for row_idx in range(2, len(validation_to_write) + 2):
                    cell = validation_sheet.cell(row=row_idx, column=col_idx)
                    cell.number_format = 'm/d/yy'

            max_length = 0
            column_letter = validation_sheet.cell(row=1, column=col_idx).column_letter
            header_lines = str(col_name).split('\n')
            max_length = max(max_length, max(len(line) for line in header_lines))
            if len(header_lines) > 1:
                validation_sheet.cell(row=1, column=col_idx).alignment = Alignment(wrapText=True, horizontal='center', vertical='bottom')

            for row_idx in range(2, len(validation_to_write) + 2):
                cell_value = validation_sheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    val_str = str(cell_value)
                    if "Date" in col_name:
                        val_str = "MM/DD/YYYY"
                    elif "Price" in col_name or "Total" in col_name:
                        val_str = "$#,###.00"
                    max_length = max(max_length, len(val_str))

            validation_sheet.column_dimensions[column_letter].width = (max_length + 2)

        validation_summary_df.to_excel(writer, sheet_name="Validation Summary", index=False)
        validation_summary_sheet = writer.sheets["Validation Summary"]
        for col_idx, col_name in enumerate(validation_summary_df.columns, 1):
            if col_name in {"Net Cash Impact", "Gross Cash Moved"}:
                for row_idx in range(2, len(validation_summary_df) + 2):
                    cell = validation_summary_sheet.cell(row=row_idx, column=col_idx)
                    cell.number_format = accounting_format
            max_length = len(str(col_name))
            for row_idx in range(2, len(validation_summary_df) + 2):
                cell_value = validation_summary_sheet.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    max_length = max(max_length, len(str(cell_value)))
            validation_summary_sheet.column_dimensions[
                validation_summary_sheet.cell(row=1, column=col_idx).column_letter
            ].width = max_length + 2

    print(f"Excel output saved to {output_file}")


def orders(consumer_key: str, consumer_secret: str, account_id_key: str, tokens: dict, output_file: str = None):
    """
    Main orchestration logic for fetching and processing orders.

    :param consumer_key: The E*TRADE consumer key.
    :param consumer_secret: The E*TRADE consumer secret.
    :param account_id_key: The E*TRADE account ID key.
    :param tokens: A dictionary containing the E*TRADE OAuth tokens.
    :param output_file: Optional path to the output file.
    """
    etrade_order = pyetrade.order.ETradeOrder(
        consumer_key,
        consumer_secret,
        tokens['oauth_token'],
        tokens['oauth_token_secret'],
        # dev=True  # Sandbox
        dev=False  # Production
    )

    try:
        opens, closes = fetch_executed_orders(
            etrade_order,
            account_id_key,
            from_dt=from_date,
            to_dt=to_date,
            action_mapping=action_map
        )
    except Exception as e:
        if "401" in str(e):
            print("\nError: E*TRADE API authentication failed (401 Unauthorized) while fetching orders.")
            print("Your OAuth tokens have likely expired. Please run 'python tokens.py' to generate new tokens.")
        else:
            print(f"Error fetching orders from E*TRADE: {e}")
        return

    add_expired_worthless_orders(opens, closes)

    new_trades = match_trades(opens, closes)
    
    # Bring forward historical trades from previous output
    old_trades = load_previous_output(output_file)
    
    # Merge and deduplicate
    combined = merge_and_deduplicate(old_trades, new_trades)

    write_excel_output(combined, output_file)
